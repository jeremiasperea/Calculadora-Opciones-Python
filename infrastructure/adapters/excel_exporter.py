"""Exportador a Excel."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from application.dtos.snapshot import SimulationSnapshot
from application.ports.exporter_port import ExporterPort


class ExcelExporter(ExporterPort):
    """Escribe una planilla con dos hojas: Resumen y Escenarios.

    El CSV lleva solo la curva. Aca se agrega una hoja con los supuestos y las
    metricas, que es lo que hace que la planilla se entienda sola cuando
    alguien la abre seis meses despues y no se acuerda con que volatilidad la
    corrio.

    Usa openpyxl directamente en vez de pandas. pandas ademas usa openpyxl por
    debajo para esto, asi que pasar por el dataframe agrega una dependencia
    pesada y encima quita el control sobre formato y hojas multiples.
    """

    @property
    def extension(self) -> str:
        return ".xlsx"

    @property
    def description(self) -> str:
        return "Excel (planilla con resumen y escenarios)"

    def export(self, snapshot: SimulationSnapshot, destination: Path) -> None:
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)

        libro = Workbook()
        self._escribir_resumen(libro.active, snapshot)
        self._escribir_escenarios(libro.create_sheet("Escenarios"), snapshot)
        libro.save(destination)

    def _escribir_resumen(self, hoja, snapshot: SimulationSnapshot) -> None:
        hoja.title = "Resumen"
        negrita = Font(bold=True)
        estrategia, mercado, resultado = (
            snapshot.strategy, snapshot.market, snapshot.result
        )

        fila = 1

        def seccion(titulo):
            nonlocal fila
            hoja.cell(fila, 1, titulo).font = negrita
            fila += 1

        def dato(etiqueta, valor, formato=None):
            nonlocal fila
            hoja.cell(fila, 1, etiqueta)
            celda = hoja.cell(fila, 2, valor)
            if formato:
                celda.number_format = formato
            fila += 1

        seccion("Condiciones de mercado")
        dato("Spot", mercado.spot, "#,##0.00")
        dato("Volatilidad", mercado.volatility, "0.00%")
        dato("Tasa", mercado.rate, "0.00%")
        dato("Dividendos", mercado.dividend_yield, "0.00%")
        dato("Dias al vencimiento", mercado.days_to_expiry)
        dato("Multiplicador", estrategia.multiplier)
        fila += 1

        seccion("Patas")
        hoja.cell(fila, 1, "Tipo").font = negrita
        hoja.cell(fila, 2, "Lado").font = negrita
        hoja.cell(fila, 3, "Cantidad").font = negrita
        hoja.cell(fila, 4, "Strike").font = negrita
        hoja.cell(fila, 5, "Prima").font = negrita
        fila += 1
        for leg in estrategia.legs:
            hoja.cell(fila, 1, leg.option_type.value)
            hoja.cell(fila, 2, leg.side.value)
            hoja.cell(fila, 3, leg.quantity)
            hoja.cell(fila, 4, leg.strike)
            hoja.cell(fila, 5, leg.premium)
            fila += 1
        fila += 1

        seccion("Resultado")
        dato("P&L inicial", resultado.net_premium, "#,##0.00")
        dato("Ganancia maxima", resultado.max_pnl, "#,##0.00")
        dato("Perdida maxima", resultado.min_pnl, "#,##0.00")
        dato("Break-even", ", ".join(f"{b:.2f}" for b in resultado.breakevens) or "—")
        dato("Probabilidad de beneficio", resultado.profit_probability, "0.00%")
        dato("P&L esperado", resultado.expected_pnl, "#,##0.00")
        fila += 1

        seccion("Griegos")
        for etiqueta, valor in [
            ("Delta", resultado.greeks.delta),
            ("Gamma", resultado.greeks.gamma),
            ("Vega", resultado.greeks.vega),
            ("Theta", resultado.greeks.theta),
            ("Rho", resultado.greeks.rho),
        ]:
            dato(etiqueta, valor, "#,##0.0000")

        hoja.column_dimensions["A"].width = 26
        hoja.column_dimensions["B"].width = 16

    def _escribir_escenarios(self, hoja, snapshot: SimulationSnapshot) -> None:
        negrita = Font(bold=True)
        centrado = Alignment(horizontal="center")

        for columna, titulo in enumerate(["Spot", "P&L"], start=1):
            celda = hoja.cell(1, columna, titulo)
            celda.font = negrita
            celda.alignment = centrado

        for i, (spot, pnl) in enumerate(
            zip(snapshot.result.prices, snapshot.result.pnl), start=2
        ):
            hoja.cell(i, 1, float(spot)).number_format = "#,##0.00"
            hoja.cell(i, 2, float(pnl)).number_format = "#,##0.00"

        hoja.column_dimensions["A"].width = 14
        hoja.column_dimensions["B"].width = 14
        hoja.freeze_panes = "A2"
