"""Tests de los exportadores.

Hay uno por formato en vez de una clase con un `if formato` adentro. Agregar
un formato nuevo es escribir una clase; las que ya andan no se tocan. Eso es
el principio abierto/cerrado en concreto: abierto a extension, cerrado a
modificacion.

CSV y JSON usan la biblioteca estandar. El codigo viejo usaba pandas para
escribir dos columnas — traer un dataframe entero para eso es cargar una
dependencia de 40 MB por algo que `csv.writer` resuelve en cuatro lineas. La
regla: no se agrega una dependencia si la stdlib ya lo hace.
"""

import csv
import json
from pathlib import Path

import pytest

from application.dtos.calculation import PriceRange
from application.dtos.snapshot import SimulationSnapshot
from application.ports.exporter_port import ExporterPort
from application.use_cases.calculate_strategy import CalculateStrategyUseCase
from domain.entities.leg import Leg
from domain.entities.strategy import Strategy
from domain.value_objects.market_conditions import MarketConditions
from infrastructure.adapters.bsm_pricing import BSMPricingEngine
from infrastructure.adapters.csv_exporter import CsvExporter
from infrastructure.adapters.json_exporter import JsonExporter


@pytest.fixture
def snapshot():
    estrategia = Strategy([
        Leg("CALL", "COMPRA", 1, 1000, 40),
        Leg("CALL", "VENTA", 1, 1100, 15),
    ], multiplier=100)
    mercado = MarketConditions(spot=1000, days_to_expiry=30,
                               volatility=0.35, rate=0.05)
    resultado = CalculateStrategyUseCase(BSMPricingEngine()).execute(
        estrategia, mercado, PriceRange(0.8, 1.2, 51)
    )
    return SimulationSnapshot(strategy=estrategia, market=mercado, result=resultado)


class TestCsvExporter:
    def test_cumple_el_contrato(self):
        assert isinstance(CsvExporter(), ExporterPort)
        assert CsvExporter().extension == ".csv"

    def test_escribe_encabezado_y_filas(self, snapshot, tmp_path):
        destino = tmp_path / "escenarios.csv"
        CsvExporter().export(snapshot, destino)

        with destino.open(encoding="utf-8", newline="") as f:
            filas = list(csv.reader(f))

        assert filas[0] == ["Spot", "P&L"]
        assert len(filas) == 52  # encabezado + 51 puntos

    def test_los_numeros_coinciden_con_el_resultado(self, snapshot, tmp_path):
        destino = tmp_path / "escenarios.csv"
        CsvExporter().export(snapshot, destino)

        with destino.open(encoding="utf-8", newline="") as f:
            filas = list(csv.DictReader(f))

        assert float(filas[0]["Spot"]) == pytest.approx(snapshot.result.prices[0])
        assert float(filas[0]["P&L"]) == pytest.approx(snapshot.result.pnl[0])

    def test_crea_los_directorios_que_falten(self, snapshot, tmp_path):
        """Si el operador elige una carpeta que no existe, se crea.

        Fallar con FileNotFoundError despues de que eligio el nombre seria
        hacerle repetir el paso por algo que se resuelve solo.
        """
        destino = tmp_path / "sub" / "carpeta" / "x.csv"
        CsvExporter().export(snapshot, destino)
        assert destino.exists()


class TestJsonExporter:
    def test_cumple_el_contrato(self):
        assert isinstance(JsonExporter(), ExporterPort)
        assert JsonExporter().extension == ".json"

    def test_escribe_json_valido(self, snapshot, tmp_path):
        destino = tmp_path / "sim.json"
        JsonExporter().export(snapshot, destino)
        datos = json.loads(destino.read_text(encoding="utf-8"))
        assert isinstance(datos, dict)

    def test_incluye_numero_de_version(self, snapshot, tmp_path):
        """El formato va a cambiar; el numero permite leer archivos viejos.

        Sin version, un archivo guardado hoy es indistinguible de uno guardado
        despues de cambiar el formato, y no hay forma de migrarlo.
        """
        destino = tmp_path / "sim.json"
        JsonExporter().export(snapshot, destino)
        datos = json.loads(destino.read_text(encoding="utf-8"))
        assert datos["version"] == 1

    def test_guarda_la_estrategia_completa(self, snapshot, tmp_path):
        destino = tmp_path / "sim.json"
        JsonExporter().export(snapshot, destino)
        datos = json.loads(destino.read_text(encoding="utf-8"))

        assert datos["strategy"]["multiplier"] == 100
        patas = datos["strategy"]["legs"]
        assert len(patas) == 2
        assert patas[0] == {
            "option_type": "CALL", "side": "COMPRA",
            "quantity": 1, "strike": 1000, "premium": 40,
        }

    def test_guarda_las_condiciones_de_mercado(self, snapshot, tmp_path):
        destino = tmp_path / "sim.json"
        JsonExporter().export(snapshot, destino)
        m = json.loads(destino.read_text(encoding="utf-8"))["market"]
        assert m["spot"] == 1000
        assert m["volatility"] == 0.35
        assert m["rate"] == 0.05

    def test_guarda_las_metricas(self, snapshot, tmp_path):
        destino = tmp_path / "sim.json"
        JsonExporter().export(snapshot, destino)
        r = json.loads(destino.read_text(encoding="utf-8"))["result"]

        assert r["net_premium"] == pytest.approx(snapshot.result.net_premium)
        assert r["max_pnl"] == pytest.approx(snapshot.result.max_pnl)
        assert r["greeks"]["delta"] == pytest.approx(snapshot.result.greeks.delta)
        assert len(r["breakevens"]) == len(snapshot.result.breakevens)

    def test_el_archivo_alcanza_para_reconstruir_la_estrategia(self, snapshot, tmp_path):
        """La prueba de que el formato sirve para reimportar.

        Se reconstruye la estrategia leyendo solo el archivo y se verifica que
        produce el mismo resultado. Si faltara un dato, esto fallaria.
        """
        destino = tmp_path / "sim.json"
        JsonExporter().export(snapshot, destino)
        datos = json.loads(destino.read_text(encoding="utf-8"))

        reconstruida = Strategy(
            [Leg(p["option_type"], p["side"], p["quantity"], p["strike"], p["premium"])
             for p in datos["strategy"]["legs"]],
            multiplier=datos["strategy"]["multiplier"],
        )
        mercado = MarketConditions(**datos["market"])

        recalculado = CalculateStrategyUseCase(BSMPricingEngine()).execute(
            reconstruida, mercado, PriceRange(0.8, 1.2, 51)
        )
        assert recalculado.net_premium == pytest.approx(snapshot.result.net_premium)
        assert recalculado.greeks.delta == pytest.approx(snapshot.result.greeks.delta)

    def test_es_legible_por_una_persona(self, snapshot, tmp_path):
        """Indentado, no en una sola linea.

        Un JSON de configuracion que alguien puede tener que abrir y revisar
        se escribe indentado. El ahorro de bytes no compensa la molestia.
        """
        destino = tmp_path / "sim.json"
        JsonExporter().export(snapshot, destino)
        texto = destino.read_text(encoding="utf-8")

        # Indentado: muchas lineas, no todo en una
        assert len(texto.splitlines()) > 20
        # Y las claves de primer nivel llevan sangria
        assert '  "version"' in texto


class TestExcelExporter:
    def test_cumple_el_contrato(self):
        from infrastructure.adapters.excel_exporter import ExcelExporter

        assert isinstance(ExcelExporter(), ExporterPort)
        assert ExcelExporter().extension == ".xlsx"

    def test_escribe_dos_hojas(self, snapshot, tmp_path):
        """Escenarios para analizar, Resumen para leer.

        El CSV solo lleva la curva. Excel permite guardar tambien los
        supuestos y las metricas, que es lo que hace que la planilla se
        entienda sola dentro de seis meses.
        """
        import openpyxl
        from infrastructure.adapters.excel_exporter import ExcelExporter

        destino = tmp_path / "sim.xlsx"
        ExcelExporter().export(snapshot, destino)

        libro = openpyxl.load_workbook(destino)
        assert libro.sheetnames == ["Resumen", "Escenarios"]

    def test_la_hoja_de_escenarios_tiene_la_curva(self, snapshot, tmp_path):
        import openpyxl
        from infrastructure.adapters.excel_exporter import ExcelExporter

        destino = tmp_path / "sim.xlsx"
        ExcelExporter().export(snapshot, destino)

        hoja = openpyxl.load_workbook(destino)["Escenarios"]
        assert hoja["A1"].value == "Spot"
        assert hoja["B1"].value == "P&L"
        assert hoja.max_row == 52  # encabezado + 51 puntos
        assert hoja["A2"].value == pytest.approx(snapshot.result.prices[0])

    def test_la_hoja_de_resumen_tiene_parametros_y_metricas(self, snapshot, tmp_path):
        import openpyxl
        from infrastructure.adapters.excel_exporter import ExcelExporter

        destino = tmp_path / "sim.xlsx"
        ExcelExporter().export(snapshot, destino)

        hoja = openpyxl.load_workbook(destino)["Resumen"]
        textos = [c.value for fila in hoja.iter_rows() for c in fila if c.value]
        unido = " ".join(str(t) for t in textos)

        assert "Spot" in unido
        assert "Volatilidad" in unido
        assert "Delta" in unido
        assert "Break-even" in unido


class TestPdfExporter:
    def test_cumple_el_contrato(self):
        from infrastructure.adapters.pdf_exporter import PdfExporter

        assert isinstance(PdfExporter(), ExporterPort)
        assert PdfExporter().extension == ".pdf"

    def test_genera_un_pdf_valido(self, snapshot, tmp_path):
        from infrastructure.adapters.pdf_exporter import PdfExporter

        destino = tmp_path / "reporte.pdf"
        PdfExporter().export(snapshot, destino)

        assert destino.exists()
        # Todo PDF empieza con esta firma
        assert destino.read_bytes()[:5] == b"%PDF-"

    def test_pesa_lo_suficiente_para_tener_el_grafico(self, snapshot, tmp_path):
        """Un PDF con una imagen embebida no baja de unos kilobytes.

        Es una verificacion grosera, pero detecta el caso de que el grafico no
        se haya generado y el reporte salga con solo texto.
        """
        from infrastructure.adapters.pdf_exporter import PdfExporter

        destino = tmp_path / "reporte.pdf"
        PdfExporter().export(snapshot, destino)
        assert destino.stat().st_size > 10_000

    def test_no_deja_figuras_de_matplotlib_abiertas(self, snapshot, tmp_path):
        """Cada figura que no se cierra queda en memoria.

        En una app de escritorio que exporta muchas veces, eso es una fuga
        que termina en un aviso de matplotlib y, si se insiste, en consumo
        de memoria que no baja.
        """
        import matplotlib.pyplot as plt
        from infrastructure.adapters.pdf_exporter import PdfExporter

        antes = len(plt.get_fignums())
        PdfExporter().export(snapshot, tmp_path / "r.pdf")
        assert len(plt.get_fignums()) == antes
